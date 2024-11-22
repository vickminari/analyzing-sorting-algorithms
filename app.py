from insertion import insertion_sort
from bubble import bubble_sort
from selection import selection_sort
from bucket import bucket_sort
from shell import shell_sort
from counting import counting_sort
from radix import radixSort
import time
import statistics
from tabulate import tabulate
import csv
import pandas as pd
import os
import openpyxl

def save_to_csv(data, filename, headers):
    with open(filename, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(data)

N = 150000 #alterar para o tamanho do arquivo
FILENAME = f'numbers{N}.txt' #alterar para o arquivo desejado
REPS = 5

def read_file(filename):   #alterar para ler um arquivo de numeros
    file_path = f'lists/{filename}'
    with open(file_path, 'r', encoding='utf-8') as f:
        return [int(line.strip()) for line in f.readlines()]
    
def measure_time(algorithm, arr):
    start_time = time.time()
    algorithm(arr)
    return (time.time() - start_time) * 1000 # convertendo para milissegundos

def is_sorted(arr):
    return all(arr[i] <= arr[i + 1] for i in range(len(arr) - 1))

def main():
    algorithms = {
        "Insertion Sort": insertion_sort,
        "Bubble Sort": bubble_sort,
        "Selection Sort": selection_sort,
        "Bucket Sort": bucket_sort,
        "Shell Sort": shell_sort,
        "Counting Sort": counting_sort,
        "Radix Sort": radixSort
    }

    results = {name: [] for name in algorithms}

    for _ in range(REPS):
        arr = read_file(FILENAME)
        for name, algorithm in algorithms.items():
            duration = measure_time(algorithm, arr)
            results[name].append(duration)

    # Validação de resultados
    for name, func in algorithms.items():
        sorted_arr = arr.copy()
        func(sorted_arr)
        if not is_sorted(sorted_arr):
            print(f"Erro: {name} não ordenou corretamente!")

    # Exibir resultados detalhados
    detailed_table = []
    for name, times in results.items():
        for i, time in enumerate(times):
            detailed_table.append([name, i + 1, f"{time:.2f} ms"])

    print(tabulate(detailed_table, headers=["Algoritmo", "Execução", "Tempo (ms)"]))

    # Exibir resultados
    table = []
    for name, times in results.items():
        avg_time = statistics.mean(times)
        std_dev = statistics.stdev(times)
        table.append([name, f"{avg_time:.2f} ms", f"{std_dev:.2f} ms"])

    print(tabulate(table, headers=["Algoritmo", "Tempo Médio", "Desvio Padrão"]))

    # Criar diretórios se não existirem
    if not os.path.exists('details'):
        os.makedirs('details')
    if not os.path.exists('general'):
        os.makedirs('general')
    if not os.path.exists('results'):
        os.makedirs('results')

    
    def save_to_excel(data, filename, sheet_name, headers):
        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
        else:
            workbook = openpyxl.load_workbook(filename)
        
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(headers)
        for row in data:
            sheet.append(row)
        
        workbook.save(filename)

    # Salvar resultados detalhados em Excel
    for name, times in results.items():
        individual_data = [[i + 1, f"{time:.4f} ms"] for i, time in enumerate(times)]
        avg_time = statistics.mean(times)
        std_dev = statistics.stdev(times)
        individual_data.append(["Média", f"{avg_time:.4f} ms"])
        individual_data.append(["Desvio Padrão", f"{std_dev:.4f} ms"])
        save_to_excel(individual_data, f"results/{name.replace(' ', '_').lower()}_details.xlsx", f"N={N}", ["Execução", "Tempo (ms)"])
    
    # Salvar resultados comparativos em Excel
    comparative_data = [[name, f"{statistics.mean(times):.4f} ms"] for name, times in results.items()]
    save_to_excel(comparative_data, f"results/comparative_results.xlsx", f"N={N}", ["Algoritmo", "Tempo Médio (ms)"])

if __name__ == "__main__":
    main()