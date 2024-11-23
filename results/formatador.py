import openpyxl
import os

# Caminho do arquivo Excel original e o nome do arquivo de saída
#["bucket_sort", "counting_sort", "heap_sort", "insertion_sort", "merge_sort", "quick_sort", "radix_sort", "selection_sort"]
file_name = "bucket_sort_details.xlsx"
input_file = f"results/{file_name}"
output_file = f"results/formatados/{file_name}"

def format_excel_file(file_path, output_path):
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Iterar pelas linhas da segunda coluna (coluna B)
        for row in sheet.iter_rows(min_row=2, max_col=2, min_col=2):  # Ajuste min_row conforme necessário
            cell = row[0]
            if cell.value and isinstance(cell.value, str) and " ms" in cell.value:
                try:
                    # Remover " ms", transformar em número e formatar com 4 casas decimais
                    number = float(cell.value.replace(" ms", ""))
                    cell.value = number  # Armazenar como número
                except ValueError:
                    print(f"Erro ao processar a célula {cell.coordinate}: {cell.value}")
    
    # Criar o diretório de saída se não existir
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # Salvar o arquivo formatado em um novo arquivo
    workbook.save(output_path)
    print(f"Arquivo salvo como {output_path}")

# Chamar a função
print(os.listdir("results"))
for file_name in os.listdir("results"):
    if not file_name.endswith(".xlsx"):
        continue
    input_file = f"results/{file_name}"
    output_file = f"results/formatados/{file_name}"
    format_excel_file(input_file, output_file)